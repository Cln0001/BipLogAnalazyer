"""Writes per-log-date sheets plus an aggregated "Summary" sheet into a single
.xlsx, so a whole raid season's worth of logs can accumulate in one file.

Each per-date sheet is transposed: player names across row 2 (color-coded
by class). Players are grouped into column blocks by role (Tank / Healer /
Caster DPS / Physical DPS / Unknown), each block headed by a merged role
label in row 1 and separated by a blank column, sorted by class then name
within each group so same-class players sit together. Each block has its
own label column (not a single shared column A) and its own row count, so
a block only lists the consumable rows actually relevant to that role.

Re-running on a log whose date+code sheet already exists corrects that
sheet in place (compares old vs new consumable counts) rather than
duplicating it. The "Summary" sheet is always recomputed from every
per-date sheet currently in the workbook, showing each player's average
consumable usage *per log they attended* (not a raw sum) so missing a
raid doesn't penalize them.
"""

from __future__ import annotations

import re
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from log_analyzer.analyze import (
    ROLE_CASTER_DPS,
    ROLE_HEALER,
    ROLE_PHYSICAL_DPS,
    ROLE_TANK,
    ROLE_UNKNOWN,
    PlayerSummary,
)
from log_analyzer.consumables import NAME_CATEGORIES

TOTAL_SHEET_NAME = "Summary"
LOGS_ATTENDED_LABEL = "Logs Attended"

# WoW class colors (Classic/TBC).
CLASS_COLORS = {
    "Warrior": "C79C6E",
    "Paladin": "F58CBA",
    "Hunter": "ABD473",
    "Rogue": "FFF569",
    "Priest": "FFFFFF",
    "Shaman": "0070DE",
    "Mage": "69CCF0",
    "Warlock": "9482C9",
    "Druid": "FF7D0A",
}

ROLE_ORDER = [ROLE_TANK, ROLE_HEALER, ROLE_CASTER_DPS, ROLE_PHYSICAL_DPS, ROLE_UNKNOWN]

STAT_ROWS = ["Class", "Role"]

# Consumable rows are colored by category (not by individual name) and
# grouped into sheet sections, each headed by a full-width marker row (see
# _CATEGORY_MARKER_PREFIX). data/consumables.json's raw "category" field is
# finer-grained than what's useful on a sheet, so several raw categories
# fold into one section here.
_DISPLAY_CATEGORY = {
    "potion": "Potions",
    "mana_item": "Potions",
    "energy_item": "Potions",
    "rune": "Others",
    "explosive": "Explosives",
    "scroll": "Buffs/Items",
    "flask": "Buffs/Items",
    "elixir": "Buffs/Items",
    "bandage": "Buffs/Items",
}
# Per-name override where the raw consumables.json category (and thus the
# default display-category mapping above) doesn't match what should show on
# the sheet — "Flame Cap" is a "potion" by raw category, but belongs in
# "Others" alongside Demonic Rune/Dark Rune here, same color and everything.
_CATEGORY_OVERRIDE = {"Flame Cap": "Others"}
_CATEGORY_ORDER = ["Potions", "Others", "Explosives", "Buffs/Items", "Other"]
_CATEGORY_MARKER_PREFIX = "§CAT§"  # sentinel, stripped before display
_BLANK_ROW_LABEL = "§BLANK§"  # sentinel: an empty spacer row, inserted before every header

# Section fill color, applied to every row in that category (including its
# Total row, if any) — not the section header itself, which keeps its own
# gray marker styling.
_CATEGORY_FILL = {
    "Potions": "DAEEF3",
    "Others": "F7CDEE",
    "Explosives": "FFB3B3",
}

# Rows that get an extra bottom border regardless of category — purely
# visual section dividers within Potions/Explosives.
_BOTTOM_BORDER_ROWS = {"Mana", "Drums"}

# "Explosives/Drums" is a synthetic total row (Sapper/Nades + Drums, added
# at write time, not a real consumable.name from consumables.json) — fixed
# order so the bold total sits right below its two components.
_EXPLOSIVE_TOTAL_LABEL = "Explosives/Drums"
_EXPLOSIVE_ROW_ORDER = ["Sapper/Nades", "Drums", _EXPLOSIVE_TOTAL_LABEL]

# Generic per-category Total row (e.g. "§TOTAL§Potions", summing every other
# row in that category) — displays as the bare word "Total", bold. Distinct
# from _EXPLOSIVE_TOTAL_LABEL above, which is its own older, separately
# named mechanism; left alone rather than folded into this one.
_ROW_TOTAL_PREFIX = "§TOTAL§"
_CATEGORIES_WITH_TOTAL = {"Potions"}

# Which roles a consumable is actually relevant for — used only to blank a
# *zero* cell for a role that would never realistically use it (e.g. a tank
# showing 0 Mana potions every log). A nonzero count is never hidden, even
# for an "irrelevant" role, since that's real data, not clutter.
ROLE_RELEVANCE: dict[str, set[str]] = {
    "Mana": {ROLE_HEALER, ROLE_CASTER_DPS},
    "Mana Emerald": {ROLE_HEALER, ROLE_CASTER_DPS},
    "Demonic Rune/Dark Rune": {ROLE_HEALER, ROLE_CASTER_DPS},
    "Destro": {ROLE_CASTER_DPS},
    "Flame Cap": {ROLE_CASTER_DPS},
    "Haste": {ROLE_PHYSICAL_DPS},
    "Thistle Tea": {ROLE_PHYSICAL_DPS},
    "Ironshield Potion": {ROLE_TANK},
}


def _category_for(name: str) -> str:
    if name == _EXPLOSIVE_TOTAL_LABEL:
        return "Explosives"
    if name.startswith(_ROW_TOTAL_PREFIX):
        return name[len(_ROW_TOTAL_PREFIX):]
    if name in _CATEGORY_OVERRIDE:
        return _CATEGORY_OVERRIDE[name]
    return _DISPLAY_CATEGORY.get(NAME_CATEGORIES.get(name, ""), "Other")


def _consumable_sort_key(name: str) -> tuple[int, int, str]:
    category = _category_for(name)
    cat_rank = _CATEGORY_ORDER.index(category) if category in _CATEGORY_ORDER else len(_CATEGORY_ORDER)
    if name.startswith(_ROW_TOTAL_PREFIX):
        sub_rank = 999  # a category's Total row always sorts last within it
    elif name in _EXPLOSIVE_ROW_ORDER:
        sub_rank = _EXPLOSIVE_ROW_ORDER.index(name)
    else:
        sub_rank = -1
    return (cat_rank, sub_rank, name)


def _inject_category_totals(names: list[str]) -> list[str]:
    """Appends a `§TOTAL§<category>` row for every category in
    `_CATEGORIES_WITH_TOTAL` that has at least one of its real consumables
    present in `names` — called per role block (after role-filtering) so
    each block's Total only sums what that block actually shows.
    """
    result = list(names)
    for category in _CATEGORIES_WITH_TOTAL:
        if any(_category_for(name) == category for name in names):
            result.append(f"{_ROW_TOTAL_PREFIX}{category}")
    return result


def _font_color_for(hex_color: str) -> str:
    """Picks black or white font for readable contrast on a fill color."""
    r, g, b = int(hex_color[0:2], 16), int(hex_color[2:4], 16), int(hex_color[4:6], 16)
    luminance = (0.299 * r + 0.587 * g + 0.114 * b) / 255
    return "000000" if luminance > 0.6 else "FFFFFF"


def _ordered_groups(summaries: list[PlayerSummary]) -> list[tuple[str, list[PlayerSummary]]]:
    """Groups summaries by role (in ROLE_ORDER), dropping empty roles."""
    by_role: dict[str, list[PlayerSummary]] = {}
    for summary in summaries:
        by_role.setdefault(summary.role, []).append(summary)

    groups: list[tuple[str, list[PlayerSummary]]] = []
    for role in ROLE_ORDER:
        # Class first so same-class players sit next to each other, name as tiebreaker.
        members = sorted(by_role.get(role, []), key=lambda s: (s.class_name, s.name))
        if members:
            groups.append((role, members))
    return groups


def _used_consumable_names(summaries: list[PlayerSummary]) -> list[str]:
    names: set[str] = set()
    for summary in summaries:
        names.update(summary.consumables.keys())
    if names & {"Sapper/Nades", "Drums"}:
        names.add(_EXPLOSIVE_TOTAL_LABEL)
    return sorted(names, key=_consumable_sort_key)


def _consumables_for_role(role: str, members: list[PlayerSummary], all_names: list[str]) -> list[str]:
    """Restricts `all_names` (everything used anywhere in this report) to
    what's worth showing in this role's column block: names with no
    ROLE_RELEVANCE entry (relevant to everyone, e.g. explosives/scrolls)
    pass through unconditionally; a role-restricted name (e.g. "Mana") only
    survives if it's not relevant for `role` *and* nobody in this block
    actually has a nonzero count for it — that fallback means a genuine
    off-role usage (a hybrid spec, a one-off) still gets its own row
    instead of being silently dropped.
    """
    used_in_group = {name for summary in members for name, count in summary.consumables.items() if count}
    return [
        name
        for name in all_names
        if ROLE_RELEVANCE.get(name) is None or role in ROLE_RELEVANCE[name] or name in used_in_group
    ]


def _rows_with_category_markers(consumable_names: list[str]) -> list[str]:
    """Interleaves a blank spacer row + a `§CAT§<section>` marker row before
    each category's first consumable row — `consumable_names` is already
    sorted by category (via `_consumable_sort_key`), so a single pass
    suffices.
    """
    labels: list[str] = []
    last_category = None
    for name in consumable_names:
        category = _category_for(name)
        if category != last_category:
            if last_category is not None:  # no spacer above the block's very first header
                labels.append(_BLANK_ROW_LABEL)
            labels.append(f"{_CATEGORY_MARKER_PREFIX}{category}")
            last_category = category
        labels.append(name)
    return labels


def _write_sheet(
    sheet: Worksheet,
    summaries: list[PlayerSummary],
    extra_stat_rows: dict[str, dict[str, int]] | None = None,
) -> None:
    """`extra_stat_rows` (label -> {player_name: value}) are inserted between
    the Class/Role rows and the consumable rows — used by the Total sheet
    for a "Logs Attended" count that has no equivalent on a per-date sheet.

    Each role gets its own column block *and own label column* (not one
    shared column A) so a block only lists the consumable rows actually
    relevant to that role (see `_consumables_for_role`) — a Tank block never
    carries a "Mana" row, a Healer block never carries "Ironshield Potion".
    Blocks can therefore differ in row count; a player column simply runs
    out of rows below whichever stat the player doesn't have.
    """
    extra_stat_rows = extra_stat_rows or {}
    groups = _ordered_groups(summaries)
    ordered = [summary for _role, members in groups for summary in members]
    all_consumable_names = _used_consumable_names(ordered)

    name_row = 2  # row 1 is the role-group header
    first_stat_row = name_row + 1

    for stat in ("Class", "Role"):
        sheet.row_dimensions[first_stat_row + STAT_ROWS.index(stat)].hidden = True

    col_idx = 2
    max_data_row = first_stat_row
    for role, members in groups:
        label_col = col_idx
        col_idx += 1  # this block's own label column, not the global column A

        role_consumable_names = sorted(
            _inject_category_totals(_consumables_for_role(role, members, all_consumable_names)),
            key=_consumable_sort_key,
        )
        block_row_labels = STAT_ROWS + list(extra_stat_rows.keys()) + _rows_with_category_markers(role_consumable_names)
        max_data_row = max(max_data_row, first_stat_row + len(block_row_labels) - 1)

        sheet.cell(row=name_row, column=label_col, value="Name").font = Font(bold=True)
        for row_idx, label in enumerate(block_row_labels, start=first_stat_row):
            if label == _BLANK_ROW_LABEL:
                display = ""  # not None — keeps _read_sheet's "label is None ends section" scan going past it
            elif label.startswith(_CATEGORY_MARKER_PREFIX):
                category = label[len(_CATEGORY_MARKER_PREFIX):]
                display = f"⌀ {category} per Raid" if sheet.title == TOTAL_SHEET_NAME else category
            elif label.startswith(_ROW_TOTAL_PREFIX):
                display = "Total"
            else:
                display = label
            sheet.cell(row=row_idx, column=label_col, value=display).font = Font(bold=True)

        for summary in members:
            cell = sheet.cell(row=name_row, column=col_idx, value=summary.name)
            class_color = CLASS_COLORS.get(summary.class_name)
            if class_color:
                cell.fill = PatternFill(start_color=class_color, end_color=class_color, fill_type="solid")
                cell.font = Font(bold=True, color=_font_color_for(class_color))
            else:
                cell.font = Font(bold=True)

            value_by_label = {
                "Class": summary.class_name,
                "Role": summary.role,
                **{label: values.get(summary.name, 0) for label, values in extra_stat_rows.items()},
                **{name: summary.consumables.get(name, 0) for name in role_consumable_names},
            }
            if _EXPLOSIVE_TOTAL_LABEL in role_consumable_names:
                value_by_label[_EXPLOSIVE_TOTAL_LABEL] = summary.consumables.get(
                    "Sapper/Nades", 0
                ) + summary.consumables.get("Drums", 0)
            for category in _CATEGORIES_WITH_TOTAL:
                total_label = f"{_ROW_TOTAL_PREFIX}{category}"
                if total_label in role_consumable_names:
                    value_by_label[total_label] = sum(
                        summary.consumables.get(name, 0)
                        for name in role_consumable_names
                        if _category_for(name) == category and name != total_label
                    )

            for row_idx, label in enumerate(block_row_labels, start=first_stat_row):
                if label == _BLANK_ROW_LABEL or label.startswith(_CATEGORY_MARKER_PREFIX):
                    continue  # spacer row or section header row, no per-player value
                sheet.cell(row=row_idx, column=col_idx, value=value_by_label[label])

            col_idx += 1

        block_last_col = col_idx - 1
        header_cell = sheet.cell(row=1, column=label_col, value=role)
        header_cell.font = Font(bold=True)
        if block_last_col > label_col:
            sheet.merge_cells(start_row=1, start_column=label_col, end_row=1, end_column=block_last_col)

        # Highlighted/marker rows are local to this block's column range —
        # different blocks can have the same label at different row numbers
        # (e.g. a category marker may land at row 7 in one block, row 9 in
        # another), so this must run per block rather than once globally.
        for row_idx, label in enumerate(block_row_labels, start=first_stat_row):
            if label == _BLANK_ROW_LABEL:
                continue  # plain spacer row, no fill/border

            if label.startswith(_CATEGORY_MARKER_PREFIX):
                for c in range(label_col, block_last_col + 1):
                    cell = sheet.cell(row=row_idx, column=c)
                    cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
                    cell.font = Font(bold=True, italic=True)
                if block_last_col > label_col:
                    sheet.merge_cells(start_row=row_idx, start_column=label_col, end_row=row_idx, end_column=block_last_col)
                continue

            color = _CATEGORY_FILL.get(_category_for(label))
            bottom_border = Border(bottom=Side(style="thin")) if label in _BOTTOM_BORDER_ROWS else None
            if color is None and bottom_border is None:
                continue
            bold_row = label == _EXPLOSIVE_TOTAL_LABEL or label.startswith(_ROW_TOTAL_PREFIX)
            for c in range(label_col, block_last_col + 1):
                cell = sheet.cell(row=row_idx, column=c)
                if color is not None:
                    cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                cell.font = Font(bold=(bold_row or c == label_col), color="000000")
                if bottom_border is not None:
                    cell.border = bottom_border

        col_idx += 1  # blank separator column before the next group

    last_col = col_idx - 1
    for c in range(1, last_col + 1):
        max_len = 0
        for r in range(1, max_data_row + 1):
            value = sheet.cell(row=r, column=c).value
            max_len = max(max_len, len(str(value)) if value is not None else 0)
        sheet.column_dimensions[get_column_letter(c)].width = max_len + 2


_INVALID_SHEET_CHARS_RE = re.compile(r'[\\/?*\[\]:]')


def _sanitize_sheet_name(name: str) -> str:
    return _INVALID_SHEET_CHARS_RE.sub("_", name)[:31]


def _read_sheet(sheet: Worksheet) -> list[PlayerSummary]:
    """Inverse of `_write_sheet` for a plain per-date sheet (no
    `extra_stat_rows`). Each role block has its own label column (the
    column where row 2 == "Name") followed by that block's player columns,
    up to (but not including) the next block's label column — blocks can
    have different row counts, so consumable rows are read per block
    against that block's own label column rather than a shared column A.
    """
    name_row, class_row, role_row = 2, 3, 4
    first_consumable_row = role_row + 1

    label_cols = [c for c in range(2, sheet.max_column + 1) if sheet.cell(row=name_row, column=c).value == "Name"]

    summaries: list[PlayerSummary] = []
    for i, label_col in enumerate(label_cols):
        block_end = label_cols[i + 1] - 2 if i + 1 < len(label_cols) else sheet.max_column
        player_cols = range(label_col + 1, block_end + 1)

        # A blank label cell no longer reliably means "end of this block's
        # rows" — spacer rows (written as "") round-trip through openpyxl
        # save/load as None, same as a genuinely unused row past a shorter
        # block. So scan every row up to the sheet's max instead of
        # stopping at the first None, and just skip rows with no label.
        #
        # Category section header rows have no per-player data (every data
        # column in this block blank by construction) — skip them rather
        # than treating their section label as a fake consumable. Synthetic
        # total rows (Explosives/Drums, any per-category "Total") are
        # likewise skipped — they're derived at write time, never present in
        # a fresh analyze.py PlayerSummary, so keeping them would make every
        # re-run see a "changed" signature and re-flag the log as corrected
        # even when nothing actually changed. A total row's display text
        # ("Total") isn't reliably distinguishable from a real consumable
        # name, so detection uses the one thing _write_sheet sets uniquely
        # for these rows: every data cell is bold, not just the label column
        # (a normal row is only bold at the label column). Track (row,
        # label) pairs since skipped rows break the simple "row offset ==
        # index" mapping.
        consumable_rows: list[tuple[int, str]] = []
        for row in range(first_consumable_row, sheet.max_row + 1):
            label = sheet.cell(row=row, column=label_col).value
            if label is None:
                continue
            has_data = any(sheet.cell(row=row, column=c).value is not None for c in player_cols)
            is_total_row = any(sheet.cell(row=row, column=c).font.bold for c in player_cols)
            if has_data and not is_total_row:
                consumable_rows.append((row, label))

        for col in player_cols:
            name = sheet.cell(row=name_row, column=col).value
            if name is None:
                continue
            consumables = {
                label: sheet.cell(row=r, column=col).value or 0
                for r, label in consumable_rows
            }
            summaries.append(
                PlayerSummary(
                    name=name,
                    class_name=sheet.cell(row=class_row, column=col).value,
                    role=sheet.cell(row=role_row, column=col).value,
                    consumables=consumables,
                )
            )
    return summaries


def _summary_signature(summaries: list[PlayerSummary]) -> dict[str, tuple]:
    """Drop zero-count entries before comparing — a freshly analyzed summary
    only has nonzero consumables (sparse), while one read back from a sheet
    has a 0 for every consumable row on that sheet regardless of whether
    this player used it (dense, since the sheet shows every player's
    full row set). Without normalizing this, every comparison would see a
    "difference" purely from that density mismatch, never "unchanged".
    """
    return {
        s.name: (s.class_name, s.role, tuple(sorted((k, v) for k, v in s.consumables.items() if v)))
        for s in summaries
    }


def _recompute_total_sheet(workbook, date_sheet_names: list[str]) -> None:
    """Rebuilds the "Summary" sheet from every per-date sheet currently in the
    workbook: per player, the *average* consumable count across the logs
    they attended (not a raw sum), plus a "Logs Attended" count. Averaging
    over attended logs only — rather than over every log in the file —
    means missing a raid doesn't drag a player's numbers down.
    """
    logs_attended: dict[str, int] = {}
    consumable_sums: dict[str, dict[str, int]] = {}
    identity: dict[str, tuple[str, str]] = {}  # name -> (class_name, role), latest sheet wins

    for sheet_name in sorted(date_sheet_names):  # date-prefixed names sort chronologically
        for summary in _read_sheet(workbook[sheet_name]):
            logs_attended[summary.name] = logs_attended.get(summary.name, 0) + 1
            totals = consumable_sums.setdefault(summary.name, {})
            for consumable, count in summary.consumables.items():
                totals[consumable] = totals.get(consumable, 0) + count
            identity[summary.name] = (summary.class_name, summary.role)

    total_summaries = [
        PlayerSummary(
            name=name,
            class_name=class_name,
            role=role,
            consumables={
                consumable: round(total / logs_attended[name], 1)
                for consumable, total in consumable_sums.get(name, {}).items()
            },
        )
        for name, (class_name, role) in identity.items()
    ]

    if TOTAL_SHEET_NAME in workbook.sheetnames:
        del workbook[TOTAL_SHEET_NAME]
    total_sheet = workbook.create_sheet(TOTAL_SHEET_NAME, 0)
    _write_sheet(total_sheet, total_summaries, extra_stat_rows={LOGS_ATTENDED_LABEL: logs_attended})

    # Tab order: Summary first, then date sheets newest-first — date-prefixed
    # names sort chronologically, so reverse for newest-first.
    desired_order = [TOTAL_SHEET_NAME] + sorted(date_sheet_names, reverse=True)
    workbook._sheets = [workbook[name] for name in desired_order]


def sheet_already_exists(output_path: str, sheet_name: str) -> bool:
    """Cheap pre-check so guild-batch mode can skip refetching/reparsing a
    report entirely when its sheet is already in the workbook, rather than
    paying for the summary/events API calls just to find out via
    `write_report`'s own (data-fetched-already) comparison.
    """
    sheet_name = _sanitize_sheet_name(sheet_name)
    if not Path(output_path).exists():
        return False
    workbook = load_workbook(output_path, read_only=True)
    try:
        return sheet_name in workbook.sheetnames
    finally:
        workbook.close()


def write_report(summaries: list[PlayerSummary], output_path: str, sheet_name: str) -> None:
    sheet_name = _sanitize_sheet_name(sheet_name)

    if Path(output_path).exists():
        workbook = load_workbook(output_path)
    else:
        workbook = Workbook()
        workbook.active.title = TOTAL_SHEET_NAME

    if sheet_name in workbook.sheetnames:
        # Always rewrite, even if the underlying counts match — the sheet
        # may have been written by an older version of _write_sheet (a
        # layout/styling change, not a data change), and the signature
        # check below only compares values, not layout. Rewriting is local
        # and cheap (no API calls), so there's no cost to always doing it.
        existing = _read_sheet(workbook[sheet_name])
        changed = _summary_signature(existing) != _summary_signature(summaries)
        del workbook[sheet_name]
        _write_sheet(workbook.create_sheet(sheet_name), summaries)
        print(f"Log {sheet_name} {'corrected (consumable counts differed)' if changed else 'unchanged'}.")
    else:
        _write_sheet(workbook.create_sheet(sheet_name), summaries)
        print(f"Log {sheet_name} added.")

    date_sheet_names = [name for name in workbook.sheetnames if name != TOTAL_SHEET_NAME]
    _recompute_total_sheet(workbook, date_sheet_names)

    workbook.save(output_path)
